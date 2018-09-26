#BeautifulSoup imports

from bs4 import BeautifulSoup as soup
#Selenium imports
from selenium import webdriver
from selenium.webdriver.support.select import Select

import time
#regex
import re
#openpyxl imports
import openpyxl
#other imports
import os


#Certified Nursing Assistant Data Obtaining Web Scraper

class Bot:

    def __init__(self):
        os.chdir('..')
        self.directory = os.getcwd()
        self.excelName = '\Broward.xlsx'

    def get_name_of_CNA(self, soup):
        headers = soup.find_all('h3')
        return headers[0].string.strip()


    def extract_addresses(self, countyIndex, num_pages):

        cell_number = 2
        wb = openpyxl.load_workbook(self.directory+self.excelName)
        sheet = wb['Sheet1']
        os.chdir(self.directory)
        sheet.cell(1, 2).value = 'Address'
        sheet.cell(1, 1).value = 'Name'

        #set up selenium browser
        driver = webdriver.Chrome("chromedriver")
        driver.get('https://appsmqa.doh.state.fl.us/MQASearchServices/HealthCareProviders')
        driver.implicitly_wait(20)
        #driver.set_window_size(300, 300)

        #Sets up the initial search parameter
        time.sleep(0.5)
        profession = Select(driver.find_element_by_id('ProfessionDD'))
        profession.select_by_index(18)
        time.sleep(0.5)
        county = Select(driver.find_element_by_id("SearchDto_County"))
        county.select_by_index(countyIndex)
        time.sleep(0.5)
        status = Select(driver.find_element_by_id("SearchDto_LicenseStatus"))
        status.select_by_index(1)
        time.sleep(0.1)
        driver.find_element_by_xpath('//*[@id="content"]/div/form/fieldset/p/input').click()
        time.sleep(2)

        first_page = True

        for page in range(1,num_pages+1):
            #Search for all links to each person
            nurses = driver.find_elements_by_partial_link_text("CNA")
            #print(len(nurses)) should always return 20

            for nurse in nurses:
                url = nurse.get_attribute('href')
                driver.execute_script("window.open('');")
                driver.switch_to_window(driver.window_handles[1])
                driver.get(url)

                #Beautiful soup starts to look further in the page
                page_soup = soup(driver.page_source, "html.parser")
                parsed_data = page_soup.find_all('dd')

                address = ''
                for line in range(5,len(parsed_data)-2):
                    #makes sure there isn't a blank line
                    if not parsed_data[line].string is None:
                        #checks to make sure the string actually has letters in it to avoid adding blank spaces
                        if re.search('[a-zA-Z]', parsed_data[line].string):
                            address += parsed_data[line].string.strip() + "\n"



                sheet.cell(cell_number, 2).value = address
                sheet.cell(cell_number,1).value = self.get_name_of_CNA(page_soup)

                wb.save('Broward.xlsx')
                if(address!=''):
                    cell_number+=1
                driver.close()
                driver.switch_to_window(driver.window_handles[0])

            print(str(page) + " page(s) scraped!")
            print(str(cell_number - 1) + " CNA address(s) collected")
            time.sleep(1)
            if(page<num_pages):
                if(first_page):
                    driver.find_element_by_xpath('//*[@id="content"]/div/div[3]/ul/li[12]/a[.="»"]').click()
                    first_page = False
                else:
                    driver.get('https://appsmqa.doh.state.fl.us/MQASearchServices/HealthCareProviders/IndexPaged?page='+str(page+1))


        print("Adresses succesfully extracted")
        driver.quit()


def startScrape(countyIndex, num_pages):

    CNADOWS = Bot()
    CNADOWS.extract_addresses(countyIndex, num_pages)

if __name__ == '__main__':
    print("CNADOWS running independently. running test mode")
    startScrape(6,2)
else:
    print("bot.py sucessfully imported")